# file created by: NOLAN AGAH

'''

GOALS:

First:
-create web scrape functions on dif files
-navigate drugs.com to find interactions between user's desired drugs

Second:
-put into spreadsheet the drug with possible interactions
-create schedule for users


'''

# File created by: NOLAN AGAH

from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import webbrowser
import tkinter as tk
from time import sleep
from tkinter import ttk
import os
from tkinter import messagebox

hours = ["{:02d}:00".format(x) for x in range(24)]
drug1 = "warfarin"
drug2 = "aspir 81"


def get_drug_names():
    global drug1
    global drug2
    drug1 = drug_entry1.get()
    drug2 = drug_entry2.get()
    print("First drug:", drug1)
    print("Second drug:", drug2)
    # You can process the drug names here as needed
    sleep(2)
    root.destroy() 

# Create the main window
root = tk.Tk()
root.title("Drug Interaction Checker")
root.geometry("600x400")

# Create frames for better organization
frame1 = tk.Frame(root, padx=20, pady=10)
frame1.pack()
frame2 = tk.Frame(root, padx=20, pady=10)
frame2.pack()
frame3 = tk.Frame(root, padx=20, pady=10)
frame3.pack()

# Create labels and entry widgets for drug inputs
drug_label1 = tk.Label(frame1, text="First drug:")
drug_label1.grid(row=0, column=0, sticky='w')
drug_entry1 = tk.Entry(frame1, width=30)
drug_entry1.grid(row=0, column=1)

drug_label2 = tk.Label(frame1, text="Second drug:")
drug_label2.grid(row=1, column=0, sticky='w')
drug_entry2 = tk.Entry(frame1, width=30)
drug_entry2.grid(row=1, column=1)

# Create a button to process the drug names
submit_button = tk.Button(frame2, text="Check Interactions", command=get_drug_names)
submit_button.pack()
# spreadsheet_button = tk.Button(frame3, text="Create Spreadsheet", command=create_spreadsheet)
# spreadsheet_button.pack()

# Run the main event loop
root.mainloop()

def create_schedule(drug1_time, drug1_frequency, drug2_time, drug2_frequency):
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Schedule"

    # Add headers to the worksheet
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for index, day in enumerate(days_of_week, start=2):
        ws.cell(row=1, column=index).value = day
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 15

    for hour in range(24):
        ws.cell(row=hour + 2, column=1).value = f"{hour:02d}:00"
        ws.row_dimensions[hour + 2].height = 20

    drug1_hour = int(drug1_time.split(":")[0])
    drug2_hour = int(drug2_time.split(":")[0])

    for row in range(2, 26):
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)

            # Set the drug to be taken at the specified hours
            if row - 1 == drug1_hour:
                cell.value = drug1
            if row - 1 == drug2_hour:
                cell.value = f"{cell.value} & {drug2}" if cell.value else drug2

            # Apply a border to the cell
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.border = border

    # Save the Excel file
    file_name = "schedule.xlsx"
    wb.save(file_name)

    # Open the Excel file in the default browser
    file_url = f"file://{os.path.abspath(file_name)}"
    webbrowser.open(file_url)


def find_drug_interactions():
    driver_path = "C:/path/to/your/edgedriver.exe"
    edge_options = Options()
    edge_options.add_argument("--window-size=1920,1080")
    edge_options.add_argument("--disable-extensions")
    edge_options.add_argument("--proxy-server='direct://'")
    edge_options.add_argument("--proxy-bypass-list=*")
    edge_options.add_argument("--start-maximized")
    edge_options.add_argument('--headless')
    edge_options.add_argument('--disable-gpu')
    edge_options.add_argument('--disable-dev-shm-usage')
    edge_options.add_argument('--no-sandbox')
    edge_options.add_argument('--ignore-certificate-errors')
    service = Service(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=edge_options)
    edge_options.add_argument("--enable-javascript")
    driver.get("https://www.drugs.com/interaction/list/?drug_list=")
    elem1 = driver.find_element(By.ID, "livesearch-interaction")
    elem1.clear()
    elem1.send_keys(drug1)
    elem1.send_keys(Keys.ARROW_DOWN)
    elem1.send_keys(Keys.RETURN)
    elem2 = driver.find_element(By.ID, "livesearch-interaction")
    elem2.clear()
    elem2.send_keys(drug2)
    elem2.send_keys(Keys.ARROW_DOWN)
    elem2.send_keys(Keys.RETURN)

    # wait for the "Check Interactions" button to become clickable
    wait = WebDriverWait(driver, 10)
    check_interactions_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@class="ddc-btn" and contains(@href, "interactions-check.php")]')))
    check_interactions_button.click()

    type_interaction_major = driver.find_element(By.XPATH, '//div[@class="ddc-form-check"]//label//span')
    major_text = type_interaction_major.text

    global found
    found = False

    # Print the extracted text
    for char in major_text:
        if char.isdigit() and int(char) > 0:
            found = True
            break

    if found:
        
        parent_element = driver.find_element(By.XPATH, "//*[contains(@class, 'interactions-reference')]")

        # Find the second child <p> element using XPath
        child_elements = parent_element.find_elements(By.XPATH, ".//p")
        second_child_element = child_elements[1]  # get the second element from the list

        element_text = second_child_element.text

        print(element_text)
    else:
        print("There are no interactions found between these drugs!")

    # close the browser after you press Enter
    driver.quit()

def get_drug_names():
    global drug1
    global drug2
    drug1 = drug_entry1.get()
    drug2 = drug_entry2.get()
    print("First drug:", drug1)
    print("Second drug:", drug2)
    # You can process the drug names here as needed
    sleep(2)
    root.destroy()    


find_drug_interactions()

def create_schedule_gui():
    def submit_schedule():
        drug1_time = hour_var1.get()
        drug1_frequency = int(frequency_var1.get())
        drug2_time = hour_var2.get()
        drug2_frequency = int(frequency_var2.get())

        if drug1_time == drug2_time:
            messagebox.showerror("Error", "Drugs should not be taken at the same time.")
            return

        create_schedule(drug1_time, drug1_frequency, drug2_time, drug2_frequency)
        schedule_root.destroy()


    global date_entry
    global time_label
    global hour_var
    global frequency_entry
    global desired_tims

    schedule_root = tk.Tk()
    schedule_root.title("Create Schedule")
    schedule_root.geometry("500x300")

    frame1 = tk.Frame(schedule_root, padx=20, pady=10)
    frame1.pack()

    hour_label1 = tk.Label(frame1, text="Select hour for Drug 1:")
    hour_label1.grid(row=0, column=0, sticky='w')
    hour_var1 = tk.StringVar()
    hour_combobox1 = ttk.Combobox(frame1, values=hours, textvariable=hour_var1)
    hour_combobox1.grid(row=0, column=1)

    frequency_label1 = tk.Label(frame1, text="Frequency of Drug 1 (times/day):")
    frequency_label1.grid(row=1, column=0, sticky='w')
    frequency_var1 = tk.StringVar()
    frequency_entry1 = tk.Entry(frame1, textvariable=frequency_var1)
    frequency_entry1.grid(row=1, column=1)

    hour_label2 = tk.Label(frame1, text="Select hour for Drug 2:")
    hour_label2.grid(row=2, column=0, sticky='w')
    hour_var2 = tk.StringVar()
    hour_combobox2 = ttk.Combobox(frame1, values=hours, textvariable=hour_var2)
    hour_combobox2.grid(row=2, column=1)

    frequency_label2 = tk.Label(frame1, text="Frequency of Drug 2 (times/day):")
    frequency_label2.grid(row=3, column=0, sticky='w')
    frequency_var2 = tk.StringVar()
    frequency_entry2 = tk.Entry(frame1, textvariable=frequency_var2)
    frequency_entry2.grid(row=3, column=1)

    submit_button = tk.Button(schedule_root, text="Create Schedule", command=submit_schedule)
    submit_button.pack()

    schedule_root.mainloop()

spreadsheet_initiation = input("Would you like to create a schedule? y or n: ")
if spreadsheet_initiation == "n":
        print("That's fine! Thank you for utilizing our services today! :)")
elif spreadsheet_initiation == "y":
        create_schedule_gui()
else:
    "Thank you for utilizing our services today!"
