import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta

import tkinter as tk
from time import sleep
from tkinter import ttk
import webbrowser
import os
from tkinter import messagebox

hours = ["{:02d}:00".format(x) for x in range(24)]
drug1 = "warfarin"
drug2 = "aspir 81"

def create_schedule(drug1_time, drug1_frequency, drug2_time, drug2_frequency):
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Schedule"

    # Add headers to the worksheet
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for index, day in enumerate(days_of_week, start=2):
        ws.cell(row=1, column=index).value = day
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 15

    for hour in range(24):
        ws.cell(row=hour + 2, column=1).value = f"{hour:02d}:00"
        ws.row_dimensions[hour + 2].height = 20

    drug1_hour = int(drug1_time.split(":")[0])
    drug2_hour = int(drug2_time.split(":")[0])

    for row in range(2, 26):
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)

            # Set the drug to be taken at the specified hours
            if row - 1 == drug1_hour:
                cell.value = drug1
            if row - 1 == drug2_hour:
                cell.value = f"{cell.value} & {drug2}" if cell.value else drug2

            # Apply a border to the cell
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.border = border

    # Save the Excel file
    file_name = "schedule.xlsx"
    wb.save(file_name)

    # Open the Excel file in the default browser
    file_url = f"file://{os.path.abspath(file_name)}"
    webbrowser.open(file_url)


def create_schedule_gui():
    def submit_schedule():
        drug1_time = hour_var1.get()
        drug1_frequency = int(frequency_var1.get())
        drug2_time = hour_var2.get()
        drug2_frequency = int(frequency_var2.get())

        if drug1_time == drug2_time:
            messagebox.showerror("Error", "Drugs should not be taken at the same time.")
            return

        create_schedule(drug1_time, drug1_frequency, drug2_time, drug2_frequency)
        schedule_root.destroy()


    global date_entry
    global time_label
    global hour_var
    global frequency_entry
    global desired_tims

    schedule_root = tk.Tk()
    schedule_root.title("Create Schedule")
    schedule_root.geometry("500x300")

    frame1 = tk.Frame(schedule_root, padx=20, pady=10)
    frame1.pack()

    hour_label1 = tk.Label(frame1, text="Select hour for Drug 1:")
    hour_label1.grid(row=0, column=0, sticky='w')
    hour_var1 = tk.StringVar()
    hour_combobox1 = ttk.Combobox(frame1, values=hours, textvariable=hour_var1)
    hour_combobox1.grid(row=0, column=1)

    frequency_label1 = tk.Label(frame1, text="Frequency of Drug 1 (times/day):")
    frequency_label1.grid(row=1, column=0, sticky='w')
    frequency_var1 = tk.StringVar()
    frequency_entry1 = tk.Entry(frame1, textvariable=frequency_var1)
    frequency_entry1.grid(row=1, column=1)

    hour_label2 = tk.Label(frame1, text="Select hour for Drug 2:")
    hour_label2.grid(row=2, column=0, sticky='w')
    hour_var2 = tk.StringVar()
    hour_combobox2 = ttk.Combobox(frame1, values=hours, textvariable=hour_var2)
    hour_combobox2.grid(row=2, column=1)

    frequency_label2 = tk.Label(frame1, text="Frequency of Drug 2 (times/day):")
    frequency_label2.grid(row=3, column=0, sticky='w')
    frequency_var2 = tk.StringVar()
    frequency_entry2 = tk.Entry(frame1, textvariable=frequency_var2)
    frequency_entry2.grid(row=3, column=1)

    submit_button = tk.Button(schedule_root, text="Create Schedule", command=submit_schedule)
    submit_button.pack()

    schedule_root.mainloop()



spreadsheet_initiation = input("Would you like to create a schedule? y or n: ")
if spreadsheet_initiation == "n":
    print("That's fine! Thank you for utilizing our services today! :)")
elif spreadsheet_initiation == "y":
    create_schedule_gui()